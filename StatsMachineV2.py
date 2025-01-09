import pandas as pd 
import os, openpyxl
from datetime import datetime
from scipy.stats import f_oneway
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment

##This is where the excel files live
path = os.path.abspath(os.path.dirname(__file__))

## This is a list of all the files in said directory
file_list = [file for file in os.listdir(path) if file.endswith('.xlsx')]

##this contains all the data with in the excel files
data_frames= []

## Loops through each excel file in the file list 
for file in file_list:
    file_path = os.path.join(path, file) #gets the porper path

##This try, except block tells the user in the console if a file is not used and for what reason
    try:
        print(f"Processing file:{file_path}")
        df = pd.read_excel(file_path, engine='openpyxl')  # Specify engine
        data_frames.append(df)
    except Exception as e:
        print(f"Skipping file: {file} due to error: {e}")
    
    print(df.shape) #spits out the number of rows and columns that make up the data frame

##combines all the dataframes that are made into 1 dataframe making it continous,drops and duplicate dataframes and NAN (Not A Number) errors are replaced with 0
combined_df = pd.concat(data_frames, ignore_index=True)

combined_df.fillna(0, inplace=True) 

##This is needed for some dumb reason: inside the data frame a column called "Sample Number" is added that doesn't need to be there so this deletes the inital "Sample Number column to make room for the real on in output()"
if 'Sample Number' in combined_df.columns:
    combined_df.drop(columns=['Sample Number'], inplace=True)

##spits out some info about the data frame to the console
print(combined_df.head())
print(combined_df.info())


def output():
    #File naming and inserting a column called 'Sample Number as seen on line 39, this pushes the entire table one column the the right, and drop duplicates again
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"Consoldated Results_{timestamp}.xlsx"
    combined_df.insert(0, 'Sample Number', range(1, len(combined_df) + 1))

    #This makes a new folder to prevent data duplication and if the folder already exists then line 50 will prevent errors
    results_folder = os.path.join(path, "Consolodated_results")
    os.makedirs(results_folder, exist_ok=True)
    output_file = os.path.join(results_folder, filename)

    #Here we are initalizing the xlsxwriter object which allows us to perform advanced formatting on the excel cells
    #We also make a new worksheet called 'Summary' for the data to be written in
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Summary')
        workbook = writer.book
        worksheet = writer.sheets['Summary']

        ##These are the format variables applied
        sample_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1})
        header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1})
        cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0.00'})
        stats_format = workbook.add_format({'bold': True, 'italic': True, 'align': 'center', 'valign': 'vcenter', 'border': 2, 'bg_color': '#9BBB59', 'num_format': '0.00'})

        #this for loop will start at A1 and then loop through the entire '1' row and format it to header_format
        #enumerate means to make a coutner which counters the columns in the... something....
        for col_num, value in enumerate (combined_df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        #The top for statement is calculating the row of the excel file, the + 1 is needed to make sure we start in row 1 therefore this turns it into 1-based numbering instead of 0-based numbering
        # The nested for statement is finding the number of columns
        for row in range(1, len(combined_df) + 1):
            for col in range(len(combined_df.columns)):
                value = combined_df.iloc[row - 1, col] #this accesses the elements using iloc() this is the driver of the line of code which allows us to access specific rows and columns using thier integer postion

                #This if block is going to format the sample column with yellow
                #Maybe we can do conditional formatting with this.
                if col == 0:
                    worksheet.write(row, col, value, sample_format)
                else:
                    worksheet.write(row, col, value, cell_format)
                    

    print(f"Data successfully consoldated to {output_file}")
    return output_file


def stats(output_file):
    data = pd.read_excel(output_file, engine='openpyxl')

    #Group 1
    TotalHits = data['Total Hits']
    ATH = data['Average Hit Time for All Images'] #ATH stands for Average Total Hits

    #Group 2
    NoGoHits = data['Total No Go Hits']
    ATNGH = data['Average Hit Time for No Go Images'] #ATNGH stands for Average Time No Go Hits
    
    f_hits, p_hits = f_oneway(TotalHits, NoGoHits)
    f_times, p_times = f_oneway(ATH, ATNGH)
    
    print("ANOVA Results:")
    print(f"Total Hits vs No Go Hits: F-statistic = {f_hits}, p-value = {p_hits}")
    print(f"Average Hit Times vs Average No Go Hit Times: F-statistic = {f_times}, p-value = {p_times}")
    return f_hits, p_hits, f_times, p_times


def write_to_sheet(output_file, f_hits, p_hits, f_times, p_times):
    final_stats = [["Comparison", "F-statistic", "p-value"],
                   ["Total Hits vs. No Go Hits", f_hits, p_hits],
                   ["Average Hit Times vs. No Go Hit Times", f_times, p_times]]

    # Load the workbook using openpyxl
    workbook = load_workbook(output_file)
    worksheet = workbook["Summary"]

    # Write the final stats to the specified rows and columns
    for row_idx, row_data in enumerate(final_stats, start=9):  # Start at row 9
        for col_idx, value in enumerate(row_data, start=1):  # Start at column 1 (A)
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Define the stats formatting using openpyxl styles
    stats_style = NamedStyle(name="stats_style", font=openpyxl.styles.Font(bold=True, italic=True),
                             fill=openpyxl.styles.PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid"),
                             border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                           right=openpyxl.styles.Side(style="thin"),
                                                           top=openpyxl.styles.Side(style="thin"),
                                                           bottom=openpyxl.styles.Side(style="thin")),
                                                        alignment=openpyxl.styles.Alignment(horizontal='center')
                             )
    stats_header = NamedStyle(name="stats_header", fill = openpyxl.styles.PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type = "solid"),
                              border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                           right=openpyxl.styles.Side(style="thin"),
                                                           top=openpyxl.styles.Side(style="thin"),
                                                           bottom=openpyxl.styles.Side(style="thin")))

    
    #Apply header style
    for row in range(9,12):
        for col in range(1,4):
            worksheet.cell(row=row, column=col).style = stats_header
            
    # Apply the style to the stats box
    for row in range(10, 12):  # Rows 10-11 correspond to stats rows in Excel
        for col in range(2, 4):  # Columns B-C
            worksheet.cell(row=row, column=col).style = stats_style

    # Save the workbook
    workbook.save(output_file)
    print("Stats updated in Excel sheet.")
    




# Main program execution
output_file = output()  # Generate the initial output file
f_hits, p_hits, f_times, p_times = stats(output_file)  # Perform the statistical calculations
write_to_sheet(output_file, f_hits, p_hits, f_times, p_times)  # Write the stats to the same file

input("\nPress Enter to exit the program\nIf there is an issue scroll up")
